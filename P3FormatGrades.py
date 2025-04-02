# Aaron Garry, Mikelle Burnett, Blane Santilli, Asher Swartzberg, Henry Tuttle
# P3 - Take unorganised excel data for students and their grades and organise them by suject into different worksheets.
# Should work with any unorganised excel file 

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


# Import unorganiszed grade data 
UnorganizedData = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')
CurrSheet = UnorganizedData.active

# Create a new workbook to put the formated data on 
OrganizedData = Workbook()

# Remove the default sheet 
OrganizedData.remove(OrganizedData["Sheet"])


# Creae a class for the students 
class student:
    def __init__(self, first, last,ID, grade):
        self.first = first
        self.last = last
        self.ID = ID
        self.grade = grade
        

# function to find the next available row
def next_available_row(sheet):
    row = 2 
    # Check if cell A in the row is not empty
    while sheet[f"A{row}"].value != None: 
        row += 1  # Move to the next row if cell is filled

    # return the first row that is empty
    return row  

# Go through all the studnets in the unorganized data
for row in CurrSheet.iter_rows(min_row=2, values_only=True): # skip row 1 

    # Create a variable name for the students subject
    subject_name = row[0]  

    # Create a new sheet if there isn't one made for this students subject
    if subject_name not in OrganizedData.sheetnames:
        subject1_sheet = OrganizedData.create_sheet(subject_name)

        # Format the sheet to have collumn headers 
        subject1_sheet["A1"] = "Last Name"
        subject1_sheet["B1"] = "First Name"
        subject1_sheet["C1"] = "Student ID"
        subject1_sheet["D1"] = "Grade"
        subject1_sheet["F1"] = "Summary Statistics"
        subject1_sheet["G1"] = "Value"

        # Bold the headers 
        subject1_sheet["A1"].font = Font(bold=True)
        subject1_sheet["B1"].font = Font(bold=True)
        subject1_sheet["C1"].font = Font(bold=True)
        subject1_sheet["D1"].font = Font(bold=True)
        subject1_sheet["F1"].font = Font(bold=True)
        subject1_sheet["G1"].font = Font(bold=True)

        # Format collumns to be 5 longer then the length of the header
        a1Len = len(subject1_sheet["A1"].value) + 5
        subject1_sheet.column_dimensions["A"].width = a1Len

        b1Len = len(subject1_sheet["B1"].value) + 5
        subject1_sheet.column_dimensions["B"].width = b1Len

        c1Len = len(subject1_sheet["C1"].value) + 5
        subject1_sheet.column_dimensions["C"].width = c1Len

        d1Len = len(subject1_sheet["D1"].value) +5
        subject1_sheet.column_dimensions["D"].width = d1Len

        f1Len = len(subject1_sheet["F1"].value) + 5
        subject1_sheet.column_dimensions["F"].width = f1Len

        g1Len = len(subject1_sheet["G1"].value) + 5
        subject1_sheet.column_dimensions["G"].width = g1Len

    # Create object for student with all the attribues 
    # split the first, last name, and student ID to be different variables
    fullName = row[1]
    splitName = fullName.split("_") 
    lastName = splitName[0]
    firstName = splitName[1]
    studID = splitName[2]

    oStud = student(firstName, lastName, studID, row[2])

    # Change current sheet to what subject the object is
    CurrSheet = OrganizedData[row[0]]

    # Call next available row function to put data in an empty row
    availableRow = next_available_row(CurrSheet)

    # Add data from object to the avaialble row
    CurrSheet[f"A{availableRow}"] = oStud.last
    CurrSheet[f"B{availableRow}"] = oStud.first
    CurrSheet[f"C{availableRow}"] = oStud.ID
    CurrSheet[f"D{availableRow}"] = oStud.grade
    
# iterate through each sheet to add filters and a summary of the data 
for sheet in OrganizedData.worksheets:
    
    # Get last row and make filters for A1- D(max row)
    last_row = sheet.max_row
    sheet.auto_filter.ref = f"A1:D{last_row}"

    # Add data summaries to F1:G6
    sheet["F2"] = "Highest Grade"
    sheet["F3"] = "Lowest Grade"
    sheet["F4"] = "Mean Grade"
    sheet["F5"] = "Median Grade"
    sheet["F6"] = "Number of Students"

    sheet["G2"] = f"=MAX(D2:D{last_row})"
    sheet["G3"] = f"=MIN(D2:D{last_row})"
    sheet["G4"] = f"=AVERAGE(D2:D{last_row})"
    sheet["G5"] = f"=MEDIAN(D2:D{last_row})"
    sheet["G6"] = f"=COUNT(D2:D{last_row})"

# Save changes as a new excel file 
OrganizedData.save(filename="formatted_grades.xlsx")

#Close both workbooks
UnorganizedData.close()
OrganizedData.close()
